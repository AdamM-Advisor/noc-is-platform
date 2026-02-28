import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WARN_FILL = PatternFill(start_color="FEF3C7", fill_type="solid")
CRIT_FILL = PatternFill(start_color="FEE2E2", fill_type="solid")
GOOD_FILL = PatternFill(start_color="D1FAE5", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0'),
)


def _style_header(ws, headers, row=1):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border = THIN_BORDER


def _auto_width(ws, headers):
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def _sla_color(cell, sla):
    if sla < 85:
        cell.fill = CRIT_FILL
    elif sla < 90:
        cell.fill = WARN_FILL
    elif sla >= 95:
        cell.fill = GOOD_FILL


class ExcelExporter:

    def generate_monthly_excel(self, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "KPI per Entity"
        self._write_kpi_sheet(ws1, data.get('children_kpis', []))

        if data.get('risk_scores'):
            ws2 = wb.create_sheet("Risk Scores")
            self._write_risk_sheet(ws2, data['risk_scores'])

        if data.get('aging_tickets'):
            ws3 = wb.create_sheet("Tiket Aging")
            self._write_tickets_sheet(ws3, data['aging_tickets'])

        self._write_info_sheet(wb, data)
        return self._to_bytes(wb)

    def generate_quarterly_excel(self, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "KPI per Entity"
        self._write_kpi_sheet(ws1, data.get('children_kpis', []))

        if data.get('behavior_transition'):
            ws2 = wb.create_sheet("Behavior Transition")
            self._write_behavior_sheet(ws2, data['behavior_transition'])

        if data.get('risk_scores'):
            ws3 = wb.create_sheet("Risk Scores")
            self._write_risk_sheet(ws3, data['risk_scores'])

        self._write_info_sheet(wb, data)
        return self._to_bytes(wb)

    def generate_annual_excel(self, data):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "KPI per Entity"
        self._write_kpi_sheet(ws1, data.get('children_kpis', []))

        if data.get('behavior_transition'):
            ws2 = wb.create_sheet("Behavior Transition")
            self._write_behavior_sheet(ws2, data['behavior_transition'])

        if data.get('risk_scores'):
            ws3 = wb.create_sheet("Risk Evolution")
            self._write_risk_sheet(ws3, data['risk_scores'])

        self._write_info_sheet(wb, data)
        return self._to_bytes(wb)

    def _write_kpi_sheet(self, ws, children):
        headers = ['#', 'Entity', 'Volume', 'SLA %', 'MTTR (min)',
                   'Escalation %', 'Auto-Resolve %', 'Repeat %',
                   'Trend', 'Status', 'Behavior']
        _style_header(ws, headers)
        for i, child in enumerate(children, 2):
            ws.cell(row=i, column=1, value=i-1).border = THIN_BORDER
            ws.cell(row=i, column=2, value=child.get('name', '')).border = THIN_BORDER
            ws.cell(row=i, column=3, value=child.get('volume', 0)).border = THIN_BORDER
            sla_cell = ws.cell(row=i, column=4, value=round(child.get('sla_pct', 0), 1))
            sla_cell.border = THIN_BORDER
            _sla_color(sla_cell, child.get('sla_pct', 0))
            ws.cell(row=i, column=5, value=round(child.get('mttr', 0), 0)).border = THIN_BORDER
            ws.cell(row=i, column=6, value=round(child.get('escalation_pct', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=7, value=round(child.get('auto_resolve_pct', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=8, value=round(child.get('repeat_pct', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=9, value=child.get('trend', '')).border = THIN_BORDER
            ws.cell(row=i, column=10, value=child.get('status', '')).border = THIN_BORDER
            ws.cell(row=i, column=11, value=child.get('behavior', '')).border = THIN_BORDER
        _auto_width(ws, headers)

    def _write_risk_sheet(self, ws, risk_data):
        headers = ['#', 'Site ID', 'Site Name', 'Risk Score', 'Risk Level',
                   'SLA %', 'MTTR', 'Volume']
        _style_header(ws, headers)
        for i, r in enumerate(risk_data, 2):
            ws.cell(row=i, column=1, value=i-1).border = THIN_BORDER
            ws.cell(row=i, column=2, value=r.get('site_id', '')).border = THIN_BORDER
            ws.cell(row=i, column=3, value=r.get('site_name', '')).border = THIN_BORDER
            ws.cell(row=i, column=4, value=round(r.get('risk_score', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=5, value=r.get('risk_level', '')).border = THIN_BORDER
            ws.cell(row=i, column=6, value=round(r.get('sla_pct', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=7, value=round(r.get('mttr', 0), 0)).border = THIN_BORDER
            ws.cell(row=i, column=8, value=r.get('volume', 0)).border = THIN_BORDER
        _auto_width(ws, headers)

    def _write_tickets_sheet(self, ws, tickets):
        headers = ['#', 'Ticket ID', 'Site ID', 'Severity', 'Age (hours)',
                   'Status', 'Created At']
        _style_header(ws, headers)
        for i, t in enumerate(tickets, 2):
            ws.cell(row=i, column=1, value=i-1).border = THIN_BORDER
            ws.cell(row=i, column=2, value=t.get('ticket_id', '')).border = THIN_BORDER
            ws.cell(row=i, column=3, value=t.get('site_id', '')).border = THIN_BORDER
            ws.cell(row=i, column=4, value=t.get('severity', '')).border = THIN_BORDER
            ws.cell(row=i, column=5, value=round(t.get('age_hours', 0), 1)).border = THIN_BORDER
            ws.cell(row=i, column=6, value=t.get('status', '')).border = THIN_BORDER
            ws.cell(row=i, column=7, value=t.get('created_at', '')).border = THIN_BORDER
        _auto_width(ws, headers)

    def _write_behavior_sheet(self, ws, transitions):
        headers = ['#', 'Entity', 'Behavior Awal', 'Behavior Akhir', 'Perubahan']
        _style_header(ws, headers)
        for i, t in enumerate(transitions, 2):
            ws.cell(row=i, column=1, value=i-1).border = THIN_BORDER
            ws.cell(row=i, column=2, value=t.get('name', '')).border = THIN_BORDER
            ws.cell(row=i, column=3, value=t.get('from', '')).border = THIN_BORDER
            ws.cell(row=i, column=4, value=t.get('to', '')).border = THIN_BORDER
            ws.cell(row=i, column=5, value=t.get('change', '')).border = THIN_BORDER
        _auto_width(ws, headers)

    def _write_info_sheet(self, wb, data):
        ws = wb.create_sheet("Info")
        ws['A1'] = "NOC-IS Analytics Platform v1.0"
        ws['A2'] = f"Generated: {data.get('generated_at', datetime.now().isoformat())}"
        ws['A3'] = "Author: Dr. Adam M."
        ws['A4'] = f"Entity: {data.get('entity_name', '')} ({data.get('entity_level', '')})"
        ws['A5'] = f"Period: {data.get('period_label', '')}"
        ws.column_dimensions['A'].width = 50

    def _to_bytes(self, wb):
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()
